"""
Portfolio Tracker — portfolio.py
=================================
Reads stocks from my_portfolio.txt, fetches daily data from Yahoo Finance,
computes all metrics, updates portfolio_tracker.xlsx (Dashboard + history
sheets), and generates docs/dashboard.html for GitHub Pages.

Runs via GitHub Actions every day at 6 PM IST (12:30 PM UTC).
"""

import os
import math
import calendar
import datetime
import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import Workbook, load_workbook

# ── Paths ─────────────────────────────────────────────────────────────────────

PORTFOLIO_FILE = "my_portfolio.txt"
EXCEL_PATH     = os.path.join("data", "portfolio_tracker.xlsx")
HTML_PATH      = os.path.join("docs", "dashboard.html")

# IST timezone offset
IST = datetime.timezone(datetime.timedelta(hours=5, minutes=30))


# ── Parse config ──────────────────────────────────────────────────────────────

def parse_portfolio_file(path):
    """
    Parse my_portfolio.txt into three lists of (display_name, ticker) tuples.
    Sections: [PORTFOLIO], [PORTFOLIO_SECONDARY], [BENCHMARKS].
    Secondary stocks are always rendered after primary with a divider.
    """
    portfolio, secondary, benchmarks = [], [], []
    section = None
    with open(path) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            if line == '[PORTFOLIO]':
                section = 'portfolio'
            elif line == '[PORTFOLIO_SECONDARY]':
                section = 'secondary'
            elif line == '[BENCHMARKS]':
                section = 'benchmarks'
            elif '|' in line and section:
                name, ticker = [x.strip() for x in line.split('|', 1)]
                if section == 'portfolio':
                    portfolio.append((name, ticker))
                elif section == 'secondary':
                    secondary.append((name, ticker))
                else:
                    benchmarks.append((name, ticker))
    return portfolio, secondary, benchmarks


# ── Data fetch ────────────────────────────────────────────────────────────────

def fetch_history(ticker):
    """
    Fetch ~1 year of daily OHLCV data for a ticker.
    Returns a DataFrame indexed by date, or None on failure.
    """
    try:
        df = yf.Ticker(ticker).history(period="1y", auto_adjust=True)
        if df.empty:
            print(f"  [WARN] No data returned for {ticker}")
            return None
        df.index = pd.to_datetime(df.index).date
        return df
    except Exception as e:
        print(f"  [ERROR] {ticker}: {e}")
        return None


# ── Metric calculations ───────────────────────────────────────────────────────

def safe_round(val, decimals=2):
    """Round a value safely, returning None if it's NaN or None."""
    if val is None:
        return None
    try:
        if math.isnan(float(val)):
            return None
        return round(float(val), decimals)
    except Exception:
        return None


def safe_pct(a, b):
    """Return percentage change (a - b) / b * 100, rounded to 2dp."""
    if a is None or b is None or b == 0:
        return None
    return safe_round((a - b) / b * 100)


def rsi_14(closes):
    """Standard 14-period RSI using Wilder's smoothing (EWM)."""
    delta = closes.diff()
    gain = delta.clip(lower=0)
    loss = -delta.clip(upper=0)
    avg_gain = gain.ewm(com=13, min_periods=14).mean()
    avg_loss = loss.ewm(com=13, min_periods=14).mean()
    rs = avg_gain / avg_loss
    return (100 - 100 / (1 + rs)).iloc[-1]


def overnight_z(df, window=20):
    """
    Z Score = overnight log return / rolling std of overnight log returns.
    Overnight LR(t) = LN(Open(t) / Close(t-1))
    sigma = stdev of the previous `window` overnight log returns.
    """
    ovn_lr = np.log(df['Open'] / df['Close'].shift(1))
    sigma  = ovn_lr.shift(1).rolling(window).std()
    z = ovn_lr / sigma
    return safe_round(z.iloc[-1])


def rel_vol(volumes, window=20):
    """RelVol = today volume / average of previous `window` days' volume."""
    avg = volumes.shift(1).rolling(window).mean()
    rv = volumes / avg
    return safe_round(rv.iloc[-1])


def compute_portfolio_metrics_at(name, ticker, df, idx):
    """
    Compute all portfolio metrics using only data available up to position `idx`.
    Used both for today's dashboard and for backfilling missing history rows.
    """
    s = df.iloc[:idx + 1]
    if len(s) < 3:
        return None

    closes  = s['Close']
    volumes = s['Volume']

    today_px  = safe_round(closes.iloc[-1])
    yest_px   = safe_round(closes.iloc[-2])
    today_vol = int(volumes.iloc[-1])
    yest_vol  = int(volumes.iloc[-2])

    weekly_ret  = safe_pct(today_px, closes.iloc[-6])  if len(closes) >= 6  else None
    monthly_ret = safe_pct(today_px, closes.iloc[-22]) if len(closes) >= 22 else None

    relvol  = rel_vol(volumes)
    z_score = overnight_z(s)

    try:
        rsi = safe_round(rsi_14(closes), 1)
    except Exception:
        rsi = None

    dma200 = safe_round(closes.rolling(200).mean().iloc[-1]) if len(closes) >= 200 else None
    dma50  = safe_round(closes.rolling(50).mean().iloc[-1])  if len(closes) >= 50  else None

    yr = closes.tail(252)
    high_52w = safe_round(yr.max())
    low_52w  = safe_round(yr.min())

    return {
        'name': name, 'ticker': ticker,
        'date': s.index[-1],
        'yest_px': yest_px, 'today_px': today_px,
        'change_pct': safe_pct(today_px, yest_px),
        'weekly_ret': weekly_ret, 'monthly_ret': monthly_ret,
        'yest_vol': yest_vol, 'today_vol': today_vol,
        'relvol': relvol, 'z_score': z_score, 'rsi': rsi,
        'dma200': dma200, 'dma50': dma50,
        'high_52w': high_52w, 'low_52w': low_52w,
        'from_high': safe_pct(today_px, high_52w),
        'from_low': safe_pct(today_px, low_52w),
    }


def compute_portfolio_metrics(name, ticker, df):
    """Compute metrics for the most recent row (used for the dashboard)."""
    if df is None or len(df) < 3:
        print(f"  [SKIP] Not enough data for {ticker}")
        return None
    m = compute_portfolio_metrics_at(name, ticker, df, len(df) - 1)
    if m:
        m['_df'] = df  # carry raw df for backfill use in update_excel
    return m


def compute_benchmark_metrics_at(name, ticker, df, idx):
    """Compute benchmark metrics using only data up to position `idx`."""
    s = df.iloc[:idx + 1]
    if len(s) < 3:
        return None
    closes   = s['Close']
    today_px = safe_round(closes.iloc[-1])
    yest_px  = safe_round(closes.iloc[-2])
    return {
        'name': name, 'ticker': ticker,
        'date': s.index[-1],
        'yest_px': yest_px, 'today_px': today_px,
        'change_pct': safe_pct(today_px, yest_px),
        'weekly_ret':  safe_pct(today_px, closes.iloc[-6])  if len(closes) >= 6  else None,
        'monthly_ret': safe_pct(today_px, closes.iloc[-22]) if len(closes) >= 22 else None,
        'z_score': overnight_z(s),
    }


def compute_benchmark_metrics(name, ticker, df):
    """Compute benchmark metrics for the most recent row (used for the dashboard)."""
    if df is None or len(df) < 3:
        print(f"  [SKIP] Not enough data for {ticker}")
        return None
    m = compute_benchmark_metrics_at(name, ticker, df, len(df) - 1)
    if m:
        m['_df'] = df
    return m


# ── Expiry dates ──────────────────────────────────────────────────────────────

def last_weekday_of_month(year, month, weekday):
    """Return the last date in (year, month) whose weekday matches (0=Mon … 6=Sun)."""
    last_day = calendar.monthrange(year, month)[1]
    d = datetime.date(year, month, last_day)
    while d.weekday() != weekday:
        d -= datetime.timedelta(days=1)
    return d


def days_until(target, today):
    """Calendar days from today to target date."""
    return (target - today).days


def get_expiry_info(today):
    """
    Returns (is_expiry_today: bool, list_of_expiry_dicts) where each dict has:
      label    — display name
      date     — the expiry date
      days     — calendar days until expiry (0 = today)
      is_today — bool

    NSE F&O calendar:
      Stock F&O (individual stocks) → last Tuesday  of month  (monthly only)
      Nifty 50                      → last Thursday of month  (monthly)
                                    → every Thursday          (weekly)
      Bank Nifty                    → last Wednesday of month (monthly)
                                    → every Wednesday         (weekly)
    """
    # Stock F&O monthly expiry = last Tuesday of the month
    last_tue   = last_weekday_of_month(today.year, today.month, 1)
    stock_days = days_until(last_tue, today)

    # If last Tuesday already passed this month, roll to next month
    if stock_days < 0:
        nm       = today.month % 12 + 1
        ny       = today.year + (1 if today.month == 12 else 0)
        last_tue   = last_weekday_of_month(ny, nm, 1)
        stock_days = days_until(last_tue, today)

    expiries = [
        {
            'label':    'Stock F&O Expiry',
            'date':     last_tue,
            'days':     stock_days,
            'is_today': stock_days == 0,
        }
    ]

    is_expiry_today = expiries[0]['is_today']
    return is_expiry_today, expiries


# ── Excel output ──────────────────────────────────────────────────────────────

PORTFOLIO_HEADERS = [
    'Date', 'Today', 'Yesterday', 'Change %',
    'Weekly Ret %', 'Monthly Ret %', 'Today Vol', 'Yesterday Vol',
    'RelVol', 'Z Score', 'RSI', '200 DMA', '50 DMA',
    '52W High', 'From High %', '52W Low', 'From Low %',
]

BENCHMARK_HEADERS = [
    'Date', 'Today', 'Yesterday', 'Change %',
    'Weekly Ret %', 'Monthly Ret %', 'Z Score',
]


def row_from_portfolio(d):
    return [
        d['date'], d['today_px'], d['yest_px'], d['change_pct'],
        d['weekly_ret'], d['monthly_ret'], d['today_vol'], d['yest_vol'],
        d['relvol'], d['z_score'], d['rsi'], d['dma200'], d['dma50'],
        d['high_52w'], d['from_high'], d['low_52w'], d['from_low'],
    ]


def row_from_benchmark(d):
    return [
        d['date'], d['today_px'], d['yest_px'], d['change_pct'],
        d['weekly_ret'], d['monthly_ret'], d['z_score'],
    ]


def update_excel(portfolio_data, secondary_data, benchmark_data, today):
    """Rebuild the Dashboard sheet and append one row per stock to its History sheet."""
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)

    wb = load_workbook(EXCEL_PATH) if os.path.exists(EXCEL_PATH) else Workbook()

    # Remove default blank sheet if brand new
    if 'Sheet' in wb.sheetnames and len(wb.sheetnames) == 1:
        wb.remove(wb['Sheet'])

    # ── Rebuild Dashboard ─────────────────────────────────────────────────────
    if 'Dashboard' in wb.sheetnames:
        wb.remove(wb['Dashboard'])
    dash = wb.create_sheet('Dashboard', 0)

    dash.append(['PORTFOLIO'])
    dash.append(['Stock', 'Ticker'] + PORTFOLIO_HEADERS[1:])
    for d in portfolio_data:
        if d:
            dash.append([d['name'], d['ticker']] + row_from_portfolio(d)[1:])

    # Divider row for transferred account stocks
    if secondary_data:
        dash.append(['── Transferred Account ──'])
        for d in secondary_data:
            if d:
                dash.append([d['name'], d['ticker']] + row_from_portfolio(d)[1:])

    dash.append([])
    dash.append(['BENCHMARKS'])
    dash.append(['Name', 'Ticker'] + BENCHMARK_HEADERS[1:])
    for d in benchmark_data:
        if d:
            dash.append([d['name'], d['ticker']] + row_from_benchmark(d)[1:])

    # ── Append to History sheets (with backfill for missed days) ─────────────
    all_items = [(d, 'portfolio') for d in portfolio_data + secondary_data if d] + \
                [(d, 'benchmark') for d in benchmark_data if d]

    for d, kind in all_items:
        raw = d['ticker'].replace('.NS', '').replace('.BO', '')
        raw = ''.join(c for c in raw if c.isalnum() or c in ('_', '-'))
        sheet_name = (raw + '_HIST')[:31]

        if sheet_name not in wb.sheetnames:
            hs = wb.create_sheet(sheet_name)
            hs.append(PORTFOLIO_HEADERS if kind == 'portfolio' else BENCHMARK_HEADERS)
        else:
            hs = wb[sheet_name]

        # Collect dates already stored — normalise to YYYY-MM-DD
        # (openpyxl may return datetime objects like '2026-07-24 00:00:00')
        existing = {str(r[0])[:10] for r in hs.iter_rows(min_row=2, values_only=True) if r[0]}

        # Use the raw DataFrame carried in the metrics dict to backfill
        df_raw = d.get('_df')
        if df_raw is None:
            continue

        # For new stocks write only today's row.
        # For existing stocks backfill only from the last stored date onwards.
        if existing:
            last_stored = max(datetime.date.fromisoformat(s) for s in existing)
            backfill_from = last_stored
        else:
            backfill_from = today  # new stock — start from today only

        added = 0
        for i, date in enumerate(df_raw.index):
            if str(date)[:10] in existing:
                continue          # already stored
            if date < backfill_from:
                continue          # before our backfill window
            if date > today:
                continue          # don't write future dates

            if kind == 'portfolio':
                row_m = compute_portfolio_metrics_at(d['name'], d['ticker'], df_raw, i)
                if row_m:
                    hs.append(row_from_portfolio(row_m))
                    added += 1
            else:
                row_m = compute_benchmark_metrics_at(d['name'], d['ticker'], df_raw, i)
                if row_m:
                    hs.append(row_from_benchmark(row_m))
                    added += 1

        if added > 0:
            print(f"  [HIST] {sheet_name}: added {added} row(s) (backfilled missing days)")
        else:
            print(f"  [HIST] {sheet_name}: already up-to-date")

    wb.save(EXCEL_PATH)
    print(f"[EXCEL] Saved: {EXCEL_PATH}")


# ── HTML dashboard ────────────────────────────────────────────────────────────

def _td(val, suffix=''):
    """Plain table cell."""
    if val is None:
        return '<td class="na">—</td>'
    return f'<td>{val}{suffix}</td>'


def _pct(val):
    """Colour-coded percentage cell."""
    if val is None:
        return '<td class="na">—</td>'
    cls = 'pos' if val > 0 else ('neg' if val < 0 else '')
    arrow = '▲' if val > 0 else ('▼' if val < 0 else '')
    return f'<td class="{cls}">{arrow}&nbsp;{val}%</td>'


def generate_html(portfolio_data, secondary_data, benchmark_data, today, expiries, is_expiry):
    """Generate a mobile-friendly dark-theme HTML dashboard."""
    os.makedirs(os.path.dirname(HTML_PATH), exist_ok=True)

    now_ist = datetime.datetime.now(IST).strftime("%d %b %Y, %I:%M %p IST")

    # Expiry cards — one pill per instrument showing days left
    expiry_html = ''
    for e in expiries:
        if e['is_today']:
            bg    = '#e74c3c'
            badge = f"🔴 {e['label']} — TODAY"
        elif e['days'] == 1:
            bg    = '#e67e22'
            badge = f"⚠️ {e['label']} — TOMORROW"
        else:
            bg    = '#2c3e50'
            badge = f"{e['label']} — {e['days']}d ({e['date'].strftime('%d %b')})"
        expiry_html += f'<span class="badge" style="background:{bg}">{badge}</span>\n'

    def _port_row(d):
        return f"""
        <tr>
          <td class="sname">{d['name']}<br><small>{d['ticker']}</small></td>
          {_td(d['today_px'])} {_td(d['yest_px'])} {_pct(d['change_pct'])}
          {_pct(d['weekly_ret'])} {_pct(d['monthly_ret'])}
          {_td(d['today_vol'])} {_td(d['yest_vol'])}
          {_td(d['relvol'])} {_td(d['z_score'])} {_td(d['rsi'])}
          {_td(d['dma200'])} {_td(d['dma50'])}
          {_td(d['high_52w'])} {_pct(d['from_high'])}
          {_td(d['low_52w'])} {_pct(d['from_low'])}
        </tr>"""

    # Portfolio rows
    port_rows = ''
    for d in portfolio_data:
        if d:
            port_rows += _port_row(d)

    # Divider + secondary rows
    if secondary_data:
        port_rows += '''
        <tr class="divider">
          <td colspan="17">Transferred Account</td>
        </tr>'''
        for d in secondary_data:
            if d:
                port_rows += _port_row(d)

    # Benchmark rows
    bench_rows = ''
    for d in benchmark_data:
        if not d:
            continue
        bench_rows += f"""
        <tr>
          <td class="sname">{d['name']}<br><small>{d['ticker']}</small></td>
          {_td(d['today_px'])} {_td(d['yest_px'])} {_pct(d['change_pct'])}
          {_pct(d['weekly_ret'])} {_pct(d['monthly_ret'])}
          {_td(d['z_score'])}
        </tr>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Portfolio Dashboard</title>
  <style>
    *{{box-sizing:border-box;margin:0;padding:0}}
    body{{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;
         background:#0d0f14;color:#d4d4d4;padding:12px;font-size:12px}}
    h1{{font-size:17px;color:#fff;margin-bottom:2px}}
    .sub{{color:#666;font-size:11px;margin-bottom:10px}}
    .expiry-bar{{margin-bottom:14px;display:flex;flex-wrap:wrap;gap:6px}}
    .badge{{padding:3px 10px;border-radius:10px;font-size:11px;
            font-weight:700;color:#fff}}
    .sec{{font-size:11px;font-weight:700;color:#888;text-transform:uppercase;
          letter-spacing:1px;margin:16px 0 6px}}
    .wrap{{overflow-x:auto;-webkit-overflow-scrolling:touch}}
    table{{border-collapse:collapse;min-width:700px;width:100%}}
    th{{background:#161921;color:#666;font-size:10px;font-weight:600;
        text-transform:uppercase;padding:5px 8px;text-align:right;
        white-space:nowrap;border-bottom:1px solid #1e2230}}
    th:first-child{{text-align:left}}
    td{{padding:6px 8px;text-align:right;border-bottom:1px solid #161921;
       white-space:nowrap}}
    td:first-child{{text-align:left}}
    tr:hover td{{background:#13151e}}
    .sname{{font-weight:600;color:#e8e8e8}}
    .sname small{{color:#555;font-weight:400}}
    .pos{{color:#27ae60;font-weight:700}}
    .neg{{color:#e74c3c;font-weight:700}}
    .na{{color:#333}}
    .divider td{{background:#1a1d2a;color:#666;font-size:10px;font-weight:700;
                text-transform:uppercase;letter-spacing:1px;padding:4px 8px;
                border-top:1px solid #3a3f55;border-bottom:1px solid #3a3f55;
                text-align:left !important}}
    .footer{{margin-top:16px;color:#444;font-size:10px;text-align:center}}
  </style>
</head>
<body>
  <h1>Portfolio Dashboard</h1>
  <div class="sub">Updated: {now_ist}</div>

  <div class="expiry-bar">{expiry_html}</div>

  <div class="sec">Portfolio</div>
  <div class="wrap"><table>
    <thead><tr>
      <th>Stock</th><th>Today</th><th>Yesterday</th><th>Chg%</th>
      <th>Wk%</th><th>Mo%</th><th>Today Vol</th><th>Yest Vol</th>
      <th>RelVol</th><th>Z Score</th><th>RSI</th>
      <th>200 DMA</th><th>50 DMA</th><th>52W H</th><th>↓High%</th>
      <th>52W L</th><th>↑Low%</th>
    </tr></thead>
    <tbody>{port_rows}</tbody>
  </table></div>

  <div class="sec">Macro / Benchmarks</div>
  <div class="wrap"><table>
    <thead><tr>
      <th>Index / Asset</th><th>Today</th><th>Yesterday</th><th>Chg%</th>
      <th>Wk%</th><th>Mo%</th><th>Z Score</th>
    </tr></thead>
    <tbody>{bench_rows}</tbody>
  </table></div>

  <div class="footer">
    Data: Yahoo Finance &nbsp;·&nbsp; Auto-updated daily at 6 PM IST &nbsp;·&nbsp;
    <a href="portfolio_tracker.xlsx" style="color:#555">Download Excel</a>
  </div>
</body>
</html>"""

    with open(HTML_PATH, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"[HTML] Saved: {HTML_PATH}")


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("Portfolio Tracker — daily update")
    print(f"Run date: {datetime.date.today()}")
    print("=" * 60)

    if not os.path.exists(PORTFOLIO_FILE):
        print(f"[ERROR] {PORTFOLIO_FILE} not found.")
        return

    portfolio_stocks, secondary_stocks, benchmark_stocks = parse_portfolio_file(PORTFOLIO_FILE)
    print(f"\n[INFO] Portfolio  : {[t for _, t in portfolio_stocks]}")
    print(f"[INFO] Secondary  : {[t for _, t in secondary_stocks]}")
    print(f"[INFO] Benchmarks : {[t for _, t in benchmark_stocks]}")

    today = datetime.date.today()

    # ── Fetch + compute portfolio ─────────────────────────────────────────────
    portfolio_data = []
    for name, ticker in portfolio_stocks:
        print(f"\n[FETCH] {name} ({ticker})")
        df = fetch_history(ticker)
        m = compute_portfolio_metrics(name, ticker, df)
        portfolio_data.append(m)
        if m:
            print(f"  Today: ₹{m['today_px']} | Change: {m['change_pct']}%")

    secondary_data = []
    for name, ticker in secondary_stocks:
        print(f"\n[FETCH] {name} ({ticker})")
        df = fetch_history(ticker)
        m = compute_portfolio_metrics(name, ticker, df)
        secondary_data.append(m)
        if m:
            print(f"  Today: ₹{m['today_px']} | Change: {m['change_pct']}%")

    # ── Fetch + compute benchmarks ────────────────────────────────────────────
    benchmark_data = []
    for name, ticker in benchmark_stocks:
        print(f"\n[FETCH] {name} ({ticker})")
        df = fetch_history(ticker)
        m = compute_benchmark_metrics(name, ticker, df)
        benchmark_data.append(m)
        if m:
            print(f"  Today: {m['today_px']} | Change: {m['change_pct']}%")

    # ── Expiry info ───────────────────────────────────────────────────────────
    is_expiry, expiries = get_expiry_info(today)
    for e in expiries:
        tag = "TODAY" if e['is_today'] else f"{e['days']}d"
        print(f"  [EXPIRY] {e['label']} — {e['date']} ({tag})")

    # ── Write outputs ─────────────────────────────────────────────────────────
    update_excel(portfolio_data, secondary_data, benchmark_data, today)
    generate_html(portfolio_data, secondary_data, benchmark_data, today, expiries, is_expiry)

    print("\nDone.")


if __name__ == "__main__":
    main()
